import shutil

import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet

from crunner.common import AREA_IDS, EXCEL_PATH, PLOTTED_PATH, RUNS_PATH
from crunner.excel.header import HEADER_OFFSET, create_empty_row
from crunner.excel.updaters import ExcelUpdater
from crunner.excel.updaters.circuit import CircuitDistanceUpdater
from crunner.excel.updaters.plotted import PlottedDistanceUpdater
from crunner.excel.updaters.runs import RunUpdater
from crunner.excel.updaters.streets import StreetsUpdater


class ExcelManager:
    def __init__(
        self,
        updaters: list[ExcelUpdater],
        area: str = "",
        find_new_areas: bool = True,
    ):
        self._area = area
        self.updaters = updaters
        self.find_new_areas = find_new_areas

        self.workbook: xl.Workbook | None = None
        self.sheet: Worksheet | None = None
        self.area_rows: dict[str, int] = {}

    @property
    def area(self):
        return self._area

    @area.setter
    def area(self, area: str):
        self._area = area
        self.path = EXCEL_PATH / f"{area}.xlsx"

    def manage(self):
        # Area not set
        if not self._area:
            print(f"No area set to update, aborting...")
            return

        # No workbook yet
        if not self.path.exists():
            template_path = self.path.with_stem("_TEMPLATE")
            shutil.copyfile(template_path, self.path)

        # Load the XLSX sheet for the given city
        self.workbook = xl.load_workbook(self.path)
        self.sheet = self.workbook.active
        if not self.sheet:
            print(f"Could not open active sheet for {self._area}")
            return

        print(f"Updating {self.area} ({self.path.resolve()})")
        self.__evaluate_areas()

        # Update
        for updater in self.updaters:
            updater.area = self.area

            paths = list(updater.find_paths())
            print(f"{updater.name} ({len(paths)} paths)...")

            for path in paths:
                area = path.stem

                if area in self.area_rows:
                    row = self.area_rows[area]
                    updater.update(path, self.sheet, row)

        self.workbook.save(self.path)
        self.workbook.close()

    def __evaluate_areas(self):
        areas: set[str] = set()

        if not self.sheet:
            return

        rows = self.__get_rows(True)
        to_delete = set()

        for idx, row in enumerate(rows, start=HEADER_OFFSET):
            if len(row) == 0 or not (area := row[0]):
                print(f"Deleting row {idx}")
                to_delete.add(idx)
            else:
                areas.add(str(area))

        for idx in sorted(to_delete, reverse=True):
            self.sheet.delete_rows(idx)

        if self.find_new_areas:
            # Add new areas
            new_areas = self.__find_areas() - areas
            for area in new_areas:
                row = create_empty_row(area)
                print(f"Adding new area {area}")
                self.sheet.append(row)

            # Sort alphabetically
            rows = sorted(self.__get_rows(False), key=lambda row: str(row[0]))

            # Delete unsorted rows and add back sorted ones
            self.sheet.delete_rows(HEADER_OFFSET, self.sheet.max_row)
            for row in rows:
                self.sheet.append(row)

        self.area_rows = {
            str(row[0]): idx
            for idx, row in enumerate(self.__get_rows(False), start=HEADER_OFFSET)
        }

    def __get_rows(self, only_area: bool):
        return (
            self.sheet.iter_rows(
                min_row=HEADER_OFFSET,
                max_row=self.sheet.max_row,
                min_col=1,
                max_col=1 if only_area else None,
                values_only=True,
            )
            if self.sheet
            else []
        )

    def __find_areas(self) -> set[str]:
        dirs = {
            # CIRCUIT_PATH: ".json",
            PLOTTED_PATH: ".gpx",
            RUNS_PATH: ".gpx",
            # GPX_PATH: ".gpx",
            # GRAPH_PATH: ".graphml",
        }

        return {
            path.stem
            for dir, ext in dirs.items()
            for path in (dir / self.area).rglob(f"*{ext}")
        }


def update_excel():
    areas = []

    while True:
        query = f"""
Please enter the area(s) to update (separate with ,):
{"\n".join(f"\t- {prefix} ({region})" for prefix, region in AREA_IDS.items())}
or write all to update all areas 
or press Q to quit
"""
        response = input(query)
        if response.lower() == "q":
            return

        if response.lower() == "all":
            break

        if response not in AREA_IDS:
            chosen = [a.strip() for a in response.split(",")]
            if any(id in AREA_IDS for id in chosen):
                areas = {id for id in chosen if id in AREA_IDS}
                break
            else:
                print("Sorry, try again...")
                continue

        areas = [response]
        break

    # Intialize Excel manager
    updaters = [
        StreetsUpdater(response),
        CircuitDistanceUpdater(response),
        PlottedDistanceUpdater(response),
        RunUpdater(response),
    ]

    manager = ExcelManager(updaters)

    # Manage all selected areas
    for response in areas:
        manager.area = AREA_IDS[response]
        manager.manage()


def main():
    update_excel()


if __name__ == "__main__":
    main()
