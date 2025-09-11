from abc import ABC, abstractmethod
from itertools import zip_longest
from pathlib import Path
from typing import Any, Iterable, Iterator

from openpyxl.worksheet.worksheet import Worksheet

from crunner.excel.header import HEADERS, Header


def to_iterable(value: Any) -> Iterable[Any]:
    return value if isinstance(value, list) or isinstance(value, tuple) else (value,)


class ExcelUpdater(ABC):
    def __init__(
        self,
        name: str,
        area: str,
        headers: Header | list[Header],
        should_overwrite: bool | list[bool] = False,
    ):
        self.area = area
        self.headers = to_iterable(headers)
        self.name = name
        self.should_overwrite = to_iterable(should_overwrite)

        self.cols = [1 + HEADERS.index(header) for header in self.headers]

        self.path: Path
        self.row: int
        self.sheet: Worksheet

    def update(self, path: Path, sheet: Worksheet, row: int):
        self.path = path
        self.sheet = sheet
        self.row = row

        if not self.can_update():
            return

        new_values = to_iterable(self._find_new_values())

        for new_value, should_overwrite, header, col in zip_longest(
            new_values, self.should_overwrite, self.headers, self.cols
        ):
            should_overwrite = (
                should_overwrite if should_overwrite is not None else False
            )

            curr_value = self.sheet.cell(self.row, col).value
            if (
                new_value is None
                or header is None
                or (not should_overwrite and curr_value is not None)
            ):
                continue

            self.sheet.cell(self.row, col, new_value)
            print(f"\t- {self.path.stem} ({header})")

    def can_update(self):
        # Not enough rows
        if self.sheet.max_row < self.row:
            return False

        # Not enough columns in row
        row = self.sheet[self.row]
        if all(len(row) < col for col in self.cols):
            return False

        # Should not overwrite
        if all(
            not should_overwrite and row[col] is not None
            for col, should_overwrite in zip(self.cols, self.should_overwrite)
        ):
            return False

        return True

    @abstractmethod
    def find_paths(self) -> Iterator[Path]: ...

    @abstractmethod
    def _find_new_values(self) -> Iterable[Any]: ...
